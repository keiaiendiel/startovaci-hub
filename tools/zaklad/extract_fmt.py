import openpyxl, json, os, collections, re, datetime
from openpyxl.utils import get_column_letter
SCRATCH=os.environ["SCRATCH"]
NBSP=" "
wb=openpyxl.load_workbook("Záměr VPD1_5.xlsx", data_only=True)
ws=wb["Základní údaje & výpočty"]

def decimals_from(fmt):
    # look at fractional part of the number token
    m=re.search(r'[0#]\.([0#]+)', fmt or "")
    return len(m.group(1)) if m else 0

def fmt_cell(v, nf):
    if v is None: return ""
    if isinstance(v, bool): return "ANO" if v else "NE"
    if isinstance(v,(datetime.datetime,datetime.date)):
        return v.strftime("%-d.%-m.%Y")
    if isinstance(v,str): return v.strip()
    nf=nf or "General"
    is_pct="%" in nf
    if is_pct:
        dec=decimals_from(nf)
        s=f"{v*100:.{dec}f}".replace(".",",")
        return s+NBSP+"%"
    unit=None
    if "Kč" in nf: unit="Kč"
    elif "m2" in nf or "m²" in nf: unit="m²"
    if nf=="General":
        dec=0 if float(v).is_integer() else 2
    else:
        dec=decimals_from(nf)
    neg=v<0; av=abs(v)
    s=f"{av:,.{dec}f}".replace(",","§").replace(".",",").replace("§",NBSP)
    if neg: s="−"+s
    if unit: s=s+NBSP+unit
    return s

# distinct number formats
nfhist=collections.Counter()
grid={}
for row in ws.iter_rows():
    for c in row:
        if c.value is None: continue
        nf=c.number_format
        nfhist[nf]+=1
        grid[(c.row,c.column)]=fmt_cell(c.value,nf)
print("=== DISTINCT NUMBER FORMATS ===")
for nf,ct in nfhist.most_common(40):
    print(f"  {ct:5d}  {nf!r}")

# write formatted dense grid tsv
maxr=ws.max_row; maxc=ws.max_column
with open(SCRATCH+"/grid.tsv","w") as f:
    for r in range(1,maxr+1):
        f.write("\t".join(grid.get((r,c),"") for c in range(1,maxc+1))+"\n")
json.dump({f"{r},{c}":v for (r,c),v in grid.items()}, open(SCRATCH+"/grid_fmt.json","w"), ensure_ascii=False)
print("\ngrid.tsv written:",maxr,"rows x",maxc,"cols; non-empty:",len(grid))

def region(r0,r1,c0,c1,label):
    print(f"\n=== {label}  (rows {r0}-{r1}, cols {get_column_letter(c0)}-{get_column_letter(c1)}) ===")
    for r in range(r0,r1+1):
        parts=[]
        for c in range(c0,c1+1):
            v=grid.get((r,c))
            if v: parts.append(f"{get_column_letter(c)}{r}={v}")
        if parts: print("  "+" | ".join(parts))

region(2,35,2,24,"TOP-LEFT: info tabulky + plochy")
