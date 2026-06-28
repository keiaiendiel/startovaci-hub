import openpyxl, json, collections, os
from openpyxl.utils import get_column_letter
SCRATCH=os.environ["SCRATCH"]
wb=openpyxl.load_workbook("Záměr VPD1_5.xlsx", data_only=True)
ws=wb["Základní údaje & výpočty"]
print("=== SHEET 'Základní údaje & výpočty' ===")
print("max_row",ws.max_row,"max_col",ws.max_column)
merged=[str(r) for r in ws.merged_cells.ranges]
print("merged ranges:",len(merged))
colw={}
for k,d in ws.column_dimensions.items():
    if d.width: colw[k]=round(d.width,1)
# print col widths compactly to spot year-column bands (uniform widths)
print("col widths (letter:width):")
print("  "+", ".join(f"{k}:{v}" for k,v in sorted(colw.items(), key=lambda x:openpyxl.utils.column_index_from_string(x[0]))))
cells=[]; fillhist=collections.Counter(); fillsamples={}
for row in ws.iter_rows():
    for c in row:
        v=c.value
        fill=None
        try:
            if c.fill and c.fill.patternType=='solid':
                rgb=c.fill.fgColor.rgb
                if isinstance(rgb,str) and rgb not in ('00000000',): fill=rgb
        except: pass
        bold=False
        try: bold=bool(c.font and c.font.bold)
        except: pass
        fc=None
        try:
            col=c.font.color
            rgb=getattr(col,"rgb",None) if col is not None else None
            if isinstance(rgb,str) and rgb.upper().endswith("FF0000"):
                fc=rgb
        except: pass
        if v is not None or fill:
            d={"c":c.coordinate,"row":c.row,"col":c.column,"v":(str(v) if v is not None else None),"f":fill,"b":bold}
            if fc: d["fc"]=fc
            cells.append(d)
        if fill:
            fillhist[fill]+=1
            fillsamples.setdefault(fill,[])
            if len(fillsamples[fill])<6: fillsamples[fill].append(c.coordinate)
json.dump({"max_row":ws.max_row,"max_col":ws.max_column,"merged":merged,"colw":colw,"cells":cells}, open(SCRATCH+"/zaklad_dump.json","w"), ensure_ascii=False)
print("cells with content/fill:",len(cells))
print("=== FILL HISTOGRAM ===")
for col,ct in fillhist.most_common():
    print(f"  {col}  x{ct}  e.g. {fillsamples[col]}")
imgs=getattr(ws,"_images",[])
print("=== IMAGES on sheet:",len(imgs),"===")
iminfo=[]
for i,im in enumerate(imgs):
    info={"idx":i}
    try:
        a=im.anchor
        info["from"]=[a._from.col,a._from.row]
        if getattr(a,"to",None) is not None:
            info["to"]=[a.to.col,a.to.row]
    except Exception as e: info["anchor_err"]=str(e)
    for attr in ("width","height"):
        try: info[attr]=getattr(im,attr)
        except: pass
    try: info["fmt"]=im.format
    except: pass
    iminfo.append(info)
json.dump(iminfo, open(SCRATCH+"/zaklad_images.json","w"), ensure_ascii=False)
for x in iminfo: print("  img",x)
